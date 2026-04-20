"""Write award laureate chunks with acronyms to xlsx."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from .acronym import line_initials, name_initials
from .chunker import Chunk
from .folger import Passage
from .monarchs import MonarchChunk
from .monologue_archive import MonologuePassage

_CHUNK_FILL = PatternFill("solid", fgColor="E8F0FE")
_POEM_TITLE_FILL = PatternFill("solid", fgColor="FFF2CC")
_HEADER_FONT = Font(bold=True)


def write_xlsx(
    chunks: list[Chunk],
    output_path: Path,
    award_name: str = "",
    first_letter_only_from: int | None = None,
) -> None:
    """Write laureate chunks to xlsx with acronym columns."""
    wb = openpyxl.Workbook()
    _write_laureates_sheet(wb.active, chunks, first_letter_only_from)
    _write_summary_sheet(wb.create_sheet("Summary"), chunks)
    wb.save(output_path)


def _write_laureates_sheet(
    ws, chunks: list[Chunk], first_letter_only_from: int | None = None
) -> None:
    ws.title = "Laureates"
    ws.append(["Year", "Name", "Initials", "Chunk Acronym"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT

    for chunk in chunks:
        for idx, entry in enumerate(chunk.entries):
            first_only = first_letter_only_from is not None and entry.year >= first_letter_only_from
            initials = name_initials(entry.name, first_only=first_only)
            acronym_cell = chunk.acronym if idx == 0 else ""
            ws.append([entry.year, entry.name, initials, acronym_cell])
            if idx == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = _CHUNK_FILL

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 30


def write_poetry_xlsx(
    poems: list[tuple[str, list[str | None]]],
    output_path: Path,
    sheet_title: str = "Poems",
) -> None:
    """Write one or more poems to a single sheet with per-line acronyms.

    Each poem gets a bold title row followed by its lines. None entries produce
    blank rows (stanza breaks). Poems are separated by a blank row.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(["Line", "Acronym"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT

    for poem_idx, (title, lines) in enumerate(poems):
        if poem_idx > 0:
            ws.append([None, None])
        title_row = ws.max_row + 1
        ws.append([title, ""])
        for cell in ws[title_row]:
            cell.font = _HEADER_FONT
            cell.fill = _POEM_TITLE_FILL
        for line in lines:
            if line is None:
                ws.append([None, None])
            else:
                ws.append([line, line_initials(line)])

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 20
    wb.save(output_path)


def write_monarchs_xlsx(
    chunks: list[MonarchChunk],
    output_path: Path,
    subject: str = "",
) -> None:
    """Write monarch reign chunks to xlsx with transition-digit strings."""
    wb = openpyxl.Workbook()
    _write_monarchs_sheet(wb.active, chunks, subject)
    _write_monarchs_summary(wb.create_sheet("Summary"), chunks)
    wb.save(output_path)


def _write_monarchs_sheet(ws, chunks: list[MonarchChunk], subject: str) -> None:
    ws.title = subject or "Monarchs"
    ws.append(["Accession", "End", "Name", "Father", "Mother", "Last Digit", "Century String"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT

    for chunk in chunks:
        for idx, monarch in enumerate(chunk.monarchs):
            end = monarch.end_year if monarch.end_year is not None else ""
            last_digit = monarch.accession_year % 10
            century_string = chunk.transition_string if idx == 0 else ""
            ws.append([monarch.accession_year, end, monarch.name, monarch.father, monarch.mother, last_digit, century_string])
            if idx == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = _CHUNK_FILL

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 30


def _write_monarchs_summary(ws, chunks: list[MonarchChunk]) -> None:
    ws.title = "Summary"
    ws.append(["Century", "Transition String"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for chunk in chunks:
        ws.append([f"{chunk.start_year}\u2013{chunk.end_year}", chunk.transition_string])
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30


def write_shakespeare_xlsx(passages: list[Passage], output_path: Path) -> None:
    """Write Shakespeare passages to xlsx: one row per line with acronym."""
    wb = openpyxl.Workbook()
    _write_shakespeare_detail(wb.active, passages)
    _write_shakespeare_summary(wb.create_sheet("Summary"), passages)
    wb.save(output_path)


def _write_shakespeare_detail(ws, passages: list[Passage]) -> None:
    ws.title = "Passages"
    ws.append(["Play", "Character", "Passage", "Line", "Acronym"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT

    for p in passages:
        for idx, line in enumerate(p.lines):
            play = p.play_name if idx == 0 else ""
            char = p.character if idx == 0 else ""
            passage_id = p.segment_id if idx == 0 else ""
            ws.append([play, char, passage_id, line, line_initials(line)])
            if idx == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = _CHUNK_FILL

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 20


def _write_shakespeare_summary(ws, passages: list[Passage]) -> None:
    ws.title = "Summary"
    ws.append(["Play", "Character", "Passage", "Lines", "Excerpt"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for p in passages:
        ws.append([p.play_name, p.character, p.segment_id, p.line_count, p.excerpt])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 60


def write_monologue_xlsx(passages: list[MonologuePassage], output_path: Path) -> None:
    """Write Monologue Archive passages to xlsx: one row per line with acronym."""
    wb = openpyxl.Workbook()
    _write_monologue_detail(wb.active, passages)
    _write_monologue_summary(wb.create_sheet("Summary"), passages)
    wb.save(output_path)


def _write_monologue_detail(ws, passages: list[MonologuePassage]) -> None:
    ws.title = "Passages"
    ws.append(["Playwright", "Play", "Character", "Type", "Line", "Acronym"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT

    for p in passages:
        for idx, line in enumerate(p.lines):
            playwright = p.playwright if idx == 0 else ""
            play = p.play_name if idx == 0 else ""
            char = p.character if idx == 0 else ""
            ptype = p.passage_type if idx == 0 else ""
            ws.append([playwright, play, char, ptype, line, line_initials(line)])
            if idx == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = _CHUNK_FILL

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 20


def _write_monologue_summary(ws, passages: list[MonologuePassage]) -> None:
    ws.title = "Summary"
    ws.append(["Playwright", "Play", "Character", "Type", "Lines", "Excerpt"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for p in passages:
        ws.append([p.playwright, p.play_name, p.character, p.passage_type, p.line_count, p.excerpt])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 60


def _write_summary_sheet(ws, chunks: list[Chunk]) -> None:
    ws.title = "Summary"
    ws.append(["Years", "Acronym"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for chunk in chunks:
        ws.append([f"{chunk.start_year}\u2013{chunk.end_year}", chunk.acronym])
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 40
