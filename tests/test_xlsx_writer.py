import openpyxl

from wiki_acronyms.chunker import make_chunks
from wiki_acronyms.list_parser import Entry
from wiki_acronyms.xlsx_writer import write_poetry_xlsx, write_xlsx


def _chunks():
    entries = [Entry(1901, "Sully Prudhomme"), Entry(1902, "Theodor Mommsen")]
    return make_chunks(entries, chunk_years=5, chunk_start_year=1901)


def test_file_created(tmp_path):
    out = tmp_path / "test.xlsx"
    write_xlsx(_chunks(), out)
    assert out.exists()


def test_laureates_sheet_header(tmp_path):
    out = tmp_path / "test.xlsx"
    write_xlsx(_chunks(), out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Laureates"]
    header = [cell.value for cell in ws[1]]
    assert header == ["Year", "Name", "Initials", "Chunk Acronym"]


def test_laureates_sheet_first_row(tmp_path):
    out = tmp_path / "test.xlsx"
    write_xlsx(_chunks(), out)
    wb = openpyxl.load_workbook(out)
    rows = list(wb["Laureates"].iter_rows(values_only=True))
    # Header + 2 data rows
    assert len(rows) == 3
    assert rows[1] == (1901, "Sully Prudhomme", "SP", "SPTM")


def test_chunk_acronym_blank_on_non_first_row(tmp_path):
    out = tmp_path / "test.xlsx"
    write_xlsx(_chunks(), out)
    wb = openpyxl.load_workbook(out)
    rows = list(wb["Laureates"].iter_rows(values_only=True))
    assert rows[2][3] in (None, "")


def test_summary_sheet_exists(tmp_path):
    out = tmp_path / "test.xlsx"
    write_xlsx(_chunks(), out)
    wb = openpyxl.load_workbook(out)
    assert "Summary" in wb.sheetnames


def test_summary_sheet_rows(tmp_path):
    entries = [Entry(1901, "Sully Prudhomme"), Entry(1906, "Bjørnstjerne Bjørnson")]
    chunks = make_chunks(entries, chunk_years=5, chunk_start_year=1901)
    out = tmp_path / "test.xlsx"
    write_xlsx(chunks, out)
    ws = openpyxl.load_workbook(out)["Summary"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 3  # header + 2 chunks
    assert rows[1][1] == "SP"
    assert rows[2][1] == "BB"


# write_poetry_xlsx
_POEM_LINES = [
    "Shall I compare thee to a summer's day?",
    "Thou art more lovely and more temperate:",
    None,
    "So long lives this, and this gives life to thee.",
]
_POEM_LINES_2 = [
    "Rough winds do shake the darling buds of May,",
    "And summer's lease hath all too short a date:",
]


def test_poetry_file_created(tmp_path):
    out = tmp_path / "poem.xlsx"
    write_poetry_xlsx([("Test Sonnet", _POEM_LINES)], out, sheet_title="Test Sonnet")
    assert out.exists()


def test_poetry_sheet_title(tmp_path):
    out = tmp_path / "poem.xlsx"
    write_poetry_xlsx([("Test Sonnet", _POEM_LINES)], out, sheet_title="Test Sonnet")
    wb = openpyxl.load_workbook(out)
    assert "Test Sonnet" in wb.sheetnames


def test_poetry_header(tmp_path):
    out = tmp_path / "poem.xlsx"
    write_poetry_xlsx([("Test Sonnet", _POEM_LINES)], out)
    ws = openpyxl.load_workbook(out).active
    assert [c.value for c in ws[1]] == ["Line", "Acronym"]


def test_poetry_title_row(tmp_path):
    out = tmp_path / "poem.xlsx"
    write_poetry_xlsx([("Test Sonnet", _POEM_LINES)], out)
    rows = list(openpyxl.load_workbook(out).active.iter_rows(values_only=True))
    # row 0 = header, row 1 = poem title row
    assert rows[1][0] == "Test Sonnet"


def test_poetry_line_row(tmp_path):
    out = tmp_path / "poem.xlsx"
    write_poetry_xlsx([("Test Sonnet", _POEM_LINES)], out)
    rows = list(openpyxl.load_workbook(out).active.iter_rows(values_only=True))
    # row 0 = header, row 1 = title, row 2 = first line
    assert rows[2][0] == "Shall I compare thee to a summer's day?"
    assert rows[2][1] == "SICTTASD"


def test_poetry_blank_row_for_none(tmp_path):
    out = tmp_path / "poem.xlsx"
    write_poetry_xlsx([("Test Sonnet", _POEM_LINES)], out)
    rows = list(openpyxl.load_workbook(out).active.iter_rows(values_only=True))
    # header + title + 2 lines + blank + last line → blank is at index 4
    assert rows[4] == (None, None)


def test_poetry_row_count_single(tmp_path):
    out = tmp_path / "poem.xlsx"
    write_poetry_xlsx([("Test Sonnet", _POEM_LINES)], out)
    rows = list(openpyxl.load_workbook(out).active.iter_rows(values_only=True))
    # header + title + 3 text lines + 1 blank = 6
    assert len(rows) == 6


def test_poetry_two_poems_separated_by_blank(tmp_path):
    out = tmp_path / "poem.xlsx"
    write_poetry_xlsx([("Sonnet A", _POEM_LINES), ("Sonnet B", _POEM_LINES_2)], out)
    rows = list(openpyxl.load_workbook(out).active.iter_rows(values_only=True))
    # After last line of first poem there should be a blank separator row
    # header(1) + title_A(1) + 3 text + 1 blank_stanza + separator_blank(1) + title_B(1) + 2 text = 10
    assert len(rows) == 10


def test_poetry_second_poem_title_row(tmp_path):
    out = tmp_path / "poem.xlsx"
    write_poetry_xlsx([("Sonnet A", _POEM_LINES), ("Sonnet B", _POEM_LINES_2)], out)
    rows = list(openpyxl.load_workbook(out).active.iter_rows(values_only=True))
    # separator blank is at index 6, title B at index 7
    assert rows[6] == (None, None)
    assert rows[7][0] == "Sonnet B"
