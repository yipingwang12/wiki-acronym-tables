from unittest.mock import patch

import pytest

from deck_generator.monarchs import (
    Correction, Monarch, MonarchChunk, correction_years, fetch_monarchs, filter_by_accession,
    make_monarch_chunks, parse_corrections, report_imprecise_dates, stale_corrections,
)

_SAMPLE_BINDINGS = [
    {
        "person": {"value": "http://www.wikidata.org/entity/Q12345"},
        "personLabel": {"value": "Alfred the Great"},
        "start": {"value": "0871-01-01T00:00:00Z"},
        "end": {"value": "0899-01-01T00:00:00Z"},
        "fatherLabel": {"value": "Æthelwulf"},
        "motherLabel": {"value": "Osburh"},
    },
    {
        "person": {"value": "http://www.wikidata.org/entity/Q12346"},
        "personLabel": {"value": "Edward the Elder"},
        "start": {"value": "0899-01-01T00:00:00Z"},
        "end": {"value": "0924-01-01T00:00:00Z"},
        "fatherLabel": {"value": "Alfred the Great"},
        "motherLabel": {"value": "Ealhswith"},
    },
    {
        "person": {"value": "http://www.wikidata.org/entity/Q12347"},
        "personLabel": {"value": "William I"},
        "start": {"value": "1066-12-25T00:00:00Z"},
        "end": {"value": "1087-09-09T00:00:00Z"},
        "fatherLabel": {"value": "Robert I, Duke of Normandy"},
        "motherLabel": {"value": "Herleva"},
    },
]


def _mock_sparql(bindings):
    with patch("deck_generator.monarchs._sparql_session", return_value=bindings) as m:
        yield m


# fetch_monarchs
def test_fetch_basic():
    with patch("deck_generator.monarchs._sparql_session", return_value=_SAMPLE_BINDINGS):
        monarchs = fetch_monarchs(["Q18810062"])
    assert len(monarchs) == 3
    assert monarchs[0].name == "Alfred the Great"
    assert monarchs[0].accession_year == 871
    assert monarchs[0].end_year == 899
    assert monarchs[0].father == "Æthelwulf"
    assert monarchs[0].mother == "Osburh"


def test_fetch_ordered_by_accession():
    with patch("deck_generator.monarchs._sparql_session", return_value=_SAMPLE_BINDINGS):
        monarchs = fetch_monarchs(["Q18810062"])
    years = [m.accession_year for m in monarchs]
    assert years == sorted(years)


def test_fetch_missing_end_year():
    bindings = [{
        "person": {"value": "http://www.wikidata.org/entity/Q99999"},
        "personLabel": {"value": "Charles III"},
        "start": {"value": "2022-09-08T00:00:00Z"},
        "fatherLabel": {"value": "Charles, Prince of Wales"},
        "motherLabel": {"value": "Elizabeth II"},
    }]
    with patch("deck_generator.monarchs._sparql_session", return_value=bindings):
        monarchs = fetch_monarchs(["Q192444"])
    assert monarchs[0].end_year is None


def test_fetch_skips_unlabelled_q_ids():
    bindings = [
        {"person": {"value": "http://www.wikidata.org/entity/Q99"}, "personLabel": {"value": "Q99"}, "start": {"value": "0871-01-01T00:00:00Z"}},
        _SAMPLE_BINDINGS[0],
    ]
    with patch("deck_generator.monarchs._sparql_session", return_value=bindings):
        monarchs = fetch_monarchs(["Q18810062"])
    assert len(monarchs) == 1


def test_fetch_deduplicates_same_person_same_year():
    duplicate = dict(_SAMPLE_BINDINGS[0])
    with patch("deck_generator.monarchs._sparql_session", return_value=[_SAMPLE_BINDINGS[0], duplicate]):
        monarchs = fetch_monarchs(["Q18810062"])
    assert len(monarchs) == 1


def test_fetch_deduplicates_same_person_different_title():
    # Same person Q-number, two positions with different start years (title change, not new monarch)
    row_early = {
        "person": {"value": "http://www.wikidata.org/entity/Q127318"},
        "personLabel": {"value": "George III"},
        "start": {"value": "1760-01-01T00:00:00Z"},
    }
    row_late = {
        "person": {"value": "http://www.wikidata.org/entity/Q127318"},
        "personLabel": {"value": "George III"},
        "start": {"value": "1801-01-01T00:00:00Z"},
    }
    with patch("deck_generator.monarchs._sparql_session", return_value=[row_early, row_late]):
        monarchs = fetch_monarchs(["Q110324075", "Q111722535"])
    assert len(monarchs) == 1
    assert monarchs[0].accession_year == 1760  # earliest wins


def test_fetch_merges_parent_info():
    row1 = {"person": {"value": "http://www.wikidata.org/entity/Q12345"}, "personLabel": {"value": "Alfred the Great"}, "start": {"value": "0871-01-01T00:00:00Z"}}
    row2 = {**row1, "fatherLabel": {"value": "Æthelwulf"}, "motherLabel": {"value": "Osburh"}}
    with patch("deck_generator.monarchs._sparql_session", return_value=[row1, row2]):
        monarchs = fetch_monarchs(["Q18810062"])
    assert monarchs[0].father == "Æthelwulf"
    assert monarchs[0].mother == "Osburh"


def test_fetch_query_contains_position_ids():
    with patch("deck_generator.monarchs._sparql_session", return_value=[]) as mock_session:
        fetch_monarchs(["Q18810062", "Q192444"])
    query = mock_session.call_args[0][1]
    assert "wd:Q18810062" in query
    assert "wd:Q192444" in query


def test_fetch_empty():
    with patch("deck_generator.monarchs._sparql_session", return_value=[]):
        assert fetch_monarchs(["Q18810062"]) == []


def test_fetch_no_house_clause_by_default():
    with patch("deck_generator.monarchs._sparql_session", return_value=[]) as mock_session:
        fetch_monarchs(["Q18810062"])
    query = mock_session.call_args[0][1]
    assert "wdt:P53" not in query   # unfiltered position query unchanged


def test_fetch_adds_house_clause_when_given():
    with patch("deck_generator.monarchs._sparql_session", return_value=[]) as mock_session:
        fetch_monarchs(["Q268218"], house_ids=["Q5185064", "Q934262"])
    query = mock_session.call_args[0][1]
    assert "wdt:P53" in query
    assert "wd:Q5185064" in query
    assert "wd:Q934262" in query


def test_fetch_populates_wp_title():
    bindings = [{
        **_SAMPLE_BINDINGS[2],
        "wpTitle": {"value": "William I of England"},
    }]
    with patch("deck_generator.monarchs._sparql_session", return_value=bindings):
        monarchs = fetch_monarchs(["Q18810062"])
    assert monarchs[0].wp_title == "William I of England"


def test_fetch_wp_title_none_when_absent():
    with patch("deck_generator.monarchs._sparql_session", return_value=[_SAMPLE_BINDINGS[0]]):
        monarchs = fetch_monarchs(["Q18810062"])
    assert monarchs[0].wp_title is None


def test_fetch_merges_wp_title():
    # First row has no sitelink; second (same person) has one — should be saved
    row1 = {**_SAMPLE_BINDINGS[0]}
    row2 = {**_SAMPLE_BINDINGS[0], "wpTitle": {"value": "Alfred the Great"}}
    with patch("deck_generator.monarchs._sparql_session", return_value=[row1, row2]):
        monarchs = fetch_monarchs(["Q18810062"])
    assert monarchs[0].wp_title == "Alfred the Great"


# make_monarch_chunks
def _monarchs(*year_name_pairs):
    return [Monarch(name=n, accession_year=y, end_year=None, father="", mother="") for y, n in year_name_pairs]


# filter_by_accession
def test_filter_by_accession_max_caps_dynasty():
    # Abbasid case: drop the Cairo figureheads acceding after the 1258 Baghdad fall
    ms = _monarchs((750, "As-Saffah"), (1242, "Al-Musta'sim"), (1261, "Al-Mustansir (Cairo)"))
    kept = filter_by_accession(ms, max_year=1258)
    assert [m.name for m in kept] == ["As-Saffah", "Al-Musta'sim"]


def test_filter_by_accession_min_and_max_window():
    ms = _monarchs((1368, "Hongwu"), (1627, "Chongzhen"), (1644, "Southern Ming"))
    kept = filter_by_accession(ms, min_year=1368, max_year=1643)
    assert [m.name for m in kept] == ["Hongwu", "Chongzhen"]


def test_filter_by_accession_none_bounds_keep_all():
    ms = _monarchs((661, "Mu'awiya"), (744, "Marwan II"))
    assert filter_by_accession(ms) == ms


def test_chunks_single_century():
    monarchs = _monarchs((871, "Alfred"), (899, "Edward"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=800)
    assert len(chunks) == 1
    assert chunks[0].start_year == 800
    assert chunks[0].end_year == 899


def test_chunks_two_centuries():
    monarchs = _monarchs((871, "Alfred"), (924, "Æthelstan"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=800)
    assert len(chunks) == 2
    assert chunks[0].start_year == 800
    assert chunks[1].start_year == 900


def test_transition_string():
    monarchs = _monarchs((871, "Alfred"), (899, "Edward"), (924, "Æthelstan"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=800)
    assert chunks[0].transition_string == "19"
    assert chunks[1].transition_string == "4"


def test_transition_string_same_year_repeated():
    # Two monarchs acceding in 1066 → two 6s
    monarchs = _monarchs((1066, "Harold II"), (1066, "William I"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=1000)
    assert chunks[0].transition_string == "66"


def test_empty_century_skipped():
    monarchs = _monarchs((871, "Alfred"), (1066, "William I"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=800)
    years = [c.start_year for c in chunks]
    # 900–999 has no monarchs; 800–899 and 1000–1099 do
    assert 900 not in years
    assert 800 in years
    assert 1000 in years


def test_empty_monarchs():
    assert make_monarch_chunks([]) == []


# make_monarch_chunks — end-year gap fallback
def _monarchs_with_ends(*tuples):
    """Create monarchs with (accession, end, name)."""
    return [Monarch(name=n, accession_year=a, end_year=e, father="", mother="") for a, e, n in tuples]


def test_gap_inserts_end_year():
    # Edward ends 924, Æthelstan starts 927 → 924 should appear in 900s string
    monarchs = _monarchs_with_ends((899, 924, "Edward"), (927, 939, "Æthelstan"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=800)
    nineties = next(c for c in chunks if c.start_year == 900)
    assert '4' == nineties.transition_string[0]   # 924 % 10
    assert '7' == nineties.transition_string[1]   # 927 % 10


def test_no_gap_no_extra_digit():
    # Normal succession: Henry dies 1547, Edward accedes 1547 → 1547 appears once, not
    # twice. (Edward still reigning / no end year, to isolate this from the terminal-end
    # rule covered by test_dynasty_end_year_included.)
    monarchs = _monarchs_with_ends((1509, 1547, "Henry VIII"), (1547, None, "Edward VI"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=1500)
    assert chunks[0].transition_string == "97"   # 1509 % 10, 1547 % 10


def test_end_year_lands_in_correct_century():
    # Edward (899-924) is in 800s chunk; his end year 924 should appear in 900s chunk
    monarchs = _monarchs_with_ends((899, 924, "Edward"), (927, 939, "Æthelstan"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=800)
    eighties = next(c for c in chunks if c.start_year == 800)
    nineties = next(c for c in chunks if c.start_year == 900)
    assert eighties.transition_string == "9"    # only Edward's accession 899
    assert "4" in nineties.transition_string    # Edward's end 924 in 900s


def test_gap_cnut_harold(monkeypatch):
    # Cnut ends 1035, Harold Harefoot starts 1037 → 1035 should appear
    monarchs = _monarchs_with_ends((1016, 1035, "Cnut"), (1037, 1040, "Harold"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=1000)
    ts = chunks[0].transition_string
    idx_cnut = ts.index('6')     # 1016 % 10
    assert ts[idx_cnut + 1] == '5'   # 1035 % 10 inserted next
    assert ts[idx_cnut + 2] == '7'   # 1037 % 10


def test_dynasty_end_year_included():
    # Final ruler has no successor: their end year must still appear (e.g. Umayyad
    # Marwan II acceded 744, killed 750 → the dynasty-ending 750 belongs in the 700s).
    monarchs = _monarchs_with_ends((661, 680, "Mu'awiya"), (744, 750, "Marwan II"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=600)
    seven_hundreds = next(c for c in chunks if c.start_year == 700)
    assert seven_hundreds.transition_string == "40"   # 744 accession, 750 end


def test_interregnum_start_year_included():
    # Large gap between a reign's end and the next accession (a genuine interregnum):
    # the end year now marks the interregnum's start (Commonwealth: Charles I end 1649,
    # Charles II accession 1660 — 1649 must appear, unlike the old ≤5-year-only rule).
    monarchs = _monarchs_with_ends((1625, 1649, "Charles I"), (1660, 1685, "Charles II"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=1600)
    ts = chunks[0].transition_string
    assert "9" in ts        # 1649 end (interregnum start) included
    assert ts == "5905"     # 1625, 1649, 1660, 1685


def test_add_transition_year_inserts_missing_event():
    # al-Muqtadir case: Wikidata records 908–932 unbroken, omitting the brief 929
    # deposition for al-Qahir. The manual year is sorted into place.
    monarchs = _monarchs_with_ends((908, 932, "al-Muqtadir"), (932, None, "al-Qahir"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=900,
                                 add_transition_years=[929])
    assert chunks[0].transition_string == "892"   # 908, 929, 932


def test_drop_transition_year_removes_artifact():
    # al-Musta'in case: Wikidata's 865 end date is a dating artifact; the real
    # transition is 866 (al-Mu'tazz's accession), already present.
    monarchs = _monarchs_with_ends((862, 865, "al-Musta'in"), (866, None, "al-Mu'tazz"))
    assert make_monarch_chunks(monarchs, 100, 800)[0].transition_string == "256"  # 862,865,866
    chunks = make_monarch_chunks(monarchs, 100, 800, drop_transition_years=[865])
    assert chunks[0].transition_string == "26"    # 865 gone


def test_drop_removes_all_occurrences_of_year():
    monarchs = _monarchs((744, "Yazid III"), (744, "Ibrahim"), (750, "Marwan II"))
    chunks = make_monarch_chunks(monarchs, 100, 700, drop_transition_years=[744])
    assert chunks[0].transition_string == "0"


def test_add_and_drop_combined():
    monarchs = _monarchs_with_ends((900, 910, "A"), (910, 920, "B"))
    chunks = make_monarch_chunks(monarchs, 100, 900,
                                 add_transition_years=[915], drop_transition_years=[920])
    assert chunks[0].transition_string == "005"   # 900, 910, 915; 920 dropped


def test_no_end_year_no_gap_inserted():
    # Monarchs with no end_year should behave as before
    monarchs = _monarchs((871, "Alfred"), (899, "Edward"), (924, "Æthelstan"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=800)
    assert chunks[0].transition_string == "19"
    assert chunks[1].transition_string == "4"


def test_default_start_year():
    monarchs = _monarchs((1837, "Victoria"), (1901, "Edward VII"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100)
    assert chunks[0].start_year == 1837


def test_last_digit_extraction():
    monarchs = _monarchs((1900, "Edward VII"), (1910, "George V"))
    chunks = make_monarch_chunks(monarchs, chunk_years=100, chunk_start_year=1900)
    assert chunks[0].transition_string == "00"


# write_monarchs_xlsx
import openpyxl
from deck_generator.xlsx_writer import write_monarchs_xlsx


def _sample_chunks():
    m1 = Monarch("Alfred the Great", 871, 899, "Æthelwulf", "Osburh")
    m2 = Monarch("Edward the Elder", 899, 924, "Alfred the Great", "Ealhswith")
    m3 = Monarch("William I", 1066, 1087, "Robert I", "Herleva")
    c1 = MonarchChunk(800, 899, [m1, m2], "19")
    c2 = MonarchChunk(1000, 1099, [m3], "6")
    return [c1, c2]


def test_monarchs_xlsx_created(tmp_path):
    out = tmp_path / "monarchs.xlsx"
    write_monarchs_xlsx(_sample_chunks(), out, subject="Test Monarchs")
    assert out.exists()


def test_monarchs_sheet_title(tmp_path):
    out = tmp_path / "monarchs.xlsx"
    write_monarchs_xlsx(_sample_chunks(), out, subject="Test Monarchs")
    wb = openpyxl.load_workbook(out)
    assert "Test Monarchs" in wb.sheetnames


def test_monarchs_header(tmp_path):
    out = tmp_path / "monarchs.xlsx"
    write_monarchs_xlsx(_sample_chunks(), out)
    ws = openpyxl.load_workbook(out).active
    assert ws[1][0].value == "Accession"
    assert ws[1][2].value == "Name"
    assert ws[1][6].value == "Century String"


def test_monarchs_first_data_row(tmp_path):
    out = tmp_path / "monarchs.xlsx"
    write_monarchs_xlsx(_sample_chunks(), out)
    rows = list(openpyxl.load_workbook(out).active.iter_rows(values_only=True))
    assert rows[1][0] == 871
    assert rows[1][2] == "Alfred the Great"
    assert rows[1][3] == "Æthelwulf"
    assert rows[1][5] == 1       # last digit of 871
    assert rows[1][6] == "19"    # century string on first row of chunk


def test_monarchs_century_string_blank_on_non_first(tmp_path):
    out = tmp_path / "monarchs.xlsx"
    write_monarchs_xlsx(_sample_chunks(), out)
    rows = list(openpyxl.load_workbook(out).active.iter_rows(values_only=True))
    assert rows[2][6] in (None, "")  # second monarch in chunk has no century string


def test_monarchs_summary_sheet(tmp_path):
    out = tmp_path / "monarchs.xlsx"
    write_monarchs_xlsx(_sample_chunks(), out)
    wb = openpyxl.load_workbook(out)
    assert "Summary" in wb.sheetnames
    rows = list(wb["Summary"].iter_rows(values_only=True))
    assert len(rows) == 3  # header + 2 chunks
    assert rows[1][1] == "19"
    assert rows[2][1] == "6"


# --- add idempotency: a correction must not double a year Wikidata already supplies ---
def test_add_is_idempotent_when_year_already_present():
    # The regression this guards: if Wikidata later fixes Puyi's end to 1912, a config still
    # carrying add:[1912] must not emit the digit twice.
    monarchs = _monarchs_with_ends((1908, 1912, "Puyi"))
    chunks = make_monarch_chunks(monarchs, 100, 1900, add_transition_years=[1912])
    assert chunks[0].transition_string == "82"   # 1908, 1912 — not "822"


def test_add_still_inserts_when_year_absent():
    monarchs = _monarchs_with_ends((1908, 1917, "Puyi"))
    chunks = make_monarch_chunks(monarchs, 100, 1900,
                                 drop_transition_years=[1917], add_transition_years=[1912])
    assert chunks[0].transition_string == "82"


def test_add_idempotent_against_an_accession_year():
    monarchs = _monarchs((900, "A"), (910, "B"))
    chunks = make_monarch_chunks(monarchs, 100, 900, add_transition_years=[910])
    assert chunks[0].transition_string == "00"   # 910 already an accession; not tripled


def test_same_year_accessions_still_produce_two_digits():
    # Idempotent add must not collapse genuine duplicate transition years from the data itself.
    monarchs = _monarchs((744, "Yazid III"), (744, "Ibrahim"))
    assert make_monarch_chunks(monarchs, 100, 700)[0].transition_string == "44"


# --- corrections: parsing + validation ---
def test_parse_corrections_roundtrip():
    cs = parse_corrections([
        {"year": 1446, "action": "add", "reason": "restoration", "source": "List", "checked": "2026-07-16"},
        {"year": 1917, "action": "drop", "reason": "artifact", "source": "List"},
    ])
    assert cs == [
        Correction(1446, "add", "restoration", "List", "2026-07-16"),
        Correction(1917, "drop", "artifact", "List", ""),
    ]


def test_parse_corrections_empty():
    assert parse_corrections(None) == []
    assert parse_corrections([]) == []


@pytest.mark.parametrize("entry,missing", [
    ({"action": "add", "reason": "r", "source": "s"}, "year"),
    ({"year": 1, "reason": "r", "source": "s"}, "action"),
    ({"year": 1, "action": "add", "source": "s"}, "reason"),
    ({"year": 1, "action": "add", "reason": "r"}, "source"),
])
def test_parse_corrections_requires_provenance(entry, missing):
    with pytest.raises(ValueError, match=missing):
        parse_corrections([entry])


def test_parse_corrections_rejects_bad_action():
    with pytest.raises(ValueError, match="action"):
        parse_corrections([{"year": 1, "action": "delete", "reason": "r", "source": "s"}])


def test_correction_years_splits_by_action():
    cs = parse_corrections([
        {"year": 1501, "action": "add", "reason": "r", "source": "s"},
        {"year": 1502, "action": "drop", "reason": "r", "source": "s"},
        {"year": 1729, "action": "drop", "reason": "r", "source": "s"},
    ])
    assert correction_years(cs) == ([1501], [1502, 1729])


# --- stale corrections: the point is to notice when upstream catches up ---
def test_stale_add_reported_when_wikidata_supplies_year():
    monarchs = _monarchs_with_ends((1908, 1912, "Puyi"))
    cs = [Correction(1912, "add", "r", "s")]
    assert "add 1912" in stale_corrections(monarchs, cs)[0]


def test_stale_drop_reported_when_year_gone():
    monarchs = _monarchs_with_ends((1908, 1912, "Puyi"))
    cs = [Correction(1917, "drop", "r", "s")]
    assert "drop 1917" in stale_corrections(monarchs, cs)[0]


def test_live_corrections_not_reported_stale():
    monarchs = _monarchs_with_ends((1908, 1917, "Puyi"))
    cs = [Correction(1917, "drop", "r", "s"), Correction(1912, "add", "r", "s")]
    assert stale_corrections(monarchs, cs) == []


# --- timePrecision: recorded, reported, but must NOT touch digit extraction ---
def test_report_imprecise_dates_flags_sub_year_precision():
    ms = [Monarch(name="Sigfred", accession_year=770, end_year=800, father="", mother="",
                  accession_precision=8)]
    notes = report_imprecise_dates(ms)
    assert len(notes) == 1 and "Sigfred" in notes[0] and "decade" in notes[0]


def test_report_imprecise_dates_silent_on_year_precision():
    ms = [Monarch(name="Alfred", accession_year=871, end_year=899, father="", mother="",
                  accession_precision=9),
          Monarch(name="Unknown", accession_year=900, end_year=None, father="", mother="",
                  accession_precision=None)]
    assert report_imprecise_dates(ms) == []


def test_precision_does_not_change_digits():
    # The explicit contract: a decade-precision accession still yields its literal last digit,
    # so decks are byte-identical to before the field existed.
    loose = [Monarch(name="Sigfred", accession_year=770, end_year=800, father="", mother="",
                     accession_precision=7)]
    exact = [Monarch(name="Sigfred", accession_year=770, end_year=800, father="", mother="",
                     accession_precision=11)]
    assert (make_monarch_chunks(loose, 100, 700)[0].transition_string
            == make_monarch_chunks(exact, 100, 700)[0].transition_string == "0")


def test_fetch_parses_precision_and_keeps_it_with_winning_year():
    # Dedup keeps the EARLIEST accession; precision must follow that year, not the last row seen.
    bindings = [
        {"person": {"value": "http://www.wikidata.org/entity/Q1"}, "personLabel": {"value": "Stephen"},
         "start": {"value": "1141-01-01T00:00:00Z"}, "startPrec": {"value": "11"}},
        {"person": {"value": "http://www.wikidata.org/entity/Q1"}, "personLabel": {"value": "Stephen"},
         "start": {"value": "1135-01-01T00:00:00Z"}, "startPrec": {"value": "8"}},
    ]
    with patch("deck_generator.monarchs._sparql_session", return_value=bindings):
        got = fetch_monarchs(["Q1"])
    assert got[0].accession_year == 1135
    assert got[0].accession_precision == 8   # the 1135 row's precision, not 1141's


def test_fetch_precision_absent_is_none():
    bindings = [{"person": {"value": "http://www.wikidata.org/entity/Q1"},
                 "personLabel": {"value": "X"}, "start": {"value": "1000-01-01T00:00:00Z"}}]
    with patch("deck_generator.monarchs._sparql_session", return_value=bindings):
        assert fetch_monarchs(["Q1"])[0].accession_precision is None


def test_fetch_query_requests_precision():
    with patch("deck_generator.monarchs._sparql_session", return_value=[]) as mock:
        fetch_monarchs(["Q1"])
    assert "wikibase:timePrecision" in mock.call_args[0][1]
